import datetime

from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle


class Employee:
    """Represents an Employee with basic details."""

    def __init__(self, name, dob, phone, email, designation, salary):
        """
        Initialize an Employee object.

        Args:
            name (str): Employee's name.
            dob (str): Date of birth in format 'YYYY-MM-DD'.
            phone (str): Phone number.
            email (str): Email address.
            designation (str): Employee's designation.
            salary (float): Employee's salary.
        """
        self.name = name
        self.dob = dob
        self.age = self.calculate_age()
        self.phone = phone
        self.email = email
        self.designation = designation
        self.salary = salary

    def calculate_age(self):
        """
        Calculate the age of the employee based on the date of birth.

        Returns:
            int: Age of the employee.
        """
        today = datetime.date.today()
        dob = datetime.datetime.strptime(self.dob, '%Y-%m-%d').date()
        age = today.year - dob.year - \
            ((today.month, today.day) < (dob.month, dob.day))
        return age


def validate_date(date_text):
    """
    Validate if a given string is a valid date in 'YYYY-MM-DD' format.

    Args:
        date_text (str): Date string to validate.

    Returns:S
        bool: True if the date string is valid, False otherwise.
    """
    try:
        datetime.datetime.strptime(date_text, '%Y-%m-%d')
        return True
    except ValueError:
        return False


def validate_phone(phone):
    """
    Validates a phone number.

    Args:
        phone (str): The phone number to be validated.

    Returns:
        bool: True if the phone number is valid, False otherwise.
    """
    if phone.isdigit() == False:
        print("Phone number must contain only digits.")
        return False
    if len(phone) != 10:
        print("Phone number must contain 10 digits.")
        return False
    return True


def validate_email(email):
    """
    Validates if the given email address is valid.

    Args:
        email (str): The email address to be validated.

    Returns:
        bool: True if the email is valid, False otherwise.
    """
    if "@" not in email:
        print("Email must contain '@' symbol.")
        return False
    if "." not in email:
        print("Email must contain '.' symbol.")
        return False
    return True


def show_employee_details(employee_list):
    """
    Display the details of all employees in the list.

    Args:
        employee_list (list): List of Employee objects.
    """

    if len(employee_list) == 0:
        print("No employees to display.")
        return

    for emp in employee_list:
        print(f"Name: {emp.name}")
        print(f"DOB: {emp.dob}")
        print(f"Phone: {emp.phone}")
        print(f"Email: {emp.email}")
        print(f"Age: {emp.age}")
        print(f"Designation: {emp.designation}")
        print(f"Salary: {emp.salary}")
        print("#" * 50)


def save_to_text(employee_list):
    """
    Save employee details to a text file.

    Args:
        employee_list (list): List of Employee objects.
    """
    with open('employees.txt', 'w') as file:
        for emp in employee_list:
            file.write(f"Name: {emp.name}, DOB: {emp.dob}, Phone: {emp.phone}, Email: {
                       emp.email}, Age: {emp.age}, Designation: {emp.designation}, Salary: {emp.salary}\n")


def save_to_excel(employee_list):
    """
    Save employee details to an Excel file.

    Args:
        employee_list (list): List of Employee objects.
    """
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "DOB", "Phone", "Email",
              "Age", "Designation", "Salary"])
    for emp in employee_list:
        ws.append([emp.name, emp.dob, emp.phone, emp.email,
                  emp.age, emp.designation, emp.salary])
    wb.save('employees.xlsx')


def save_to_pdf(employee_list):
    """
    Save the employee data to a PDF file.

    Args:
        employee_list (list): A list of employee objects.

    Returns:
        None
    """
    doc = SimpleDocTemplate("employees.pdf", pagesize=letter)
    elements = []

    data = [
        ["Name", "DOB", "Phone", "Email", "Age", "Designation", "Salary"]
    ]
    for emp in employee_list:
        data.append([emp.name, emp.dob, emp.phone, emp.email,
                     emp.age, emp.designation, emp.salary])

    table = Table(data)

    style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ])

    table.setStyle(style)
    elements.append(table)

    doc.build(elements)


def bulk_read_from_excel():
    """
    Read employee data from an Excel file.

    Returns:
        list: A list of Employee objects.
    """
    wb = load_workbook('employees.xlsx')
    ws = wb.active
    employee_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        employee_list.append(
            Employee(row[0], row[1], row[2], row[3], row[4], row[5]))
    return employee_list


def get_employee_details():
    """
    Prompts the user to enter employee details and returns an Employee object.

    Returns:
        Employee: An instance of the Employee class with the entered details.
    """
    name = input("Enter Name: ")
    dob = input("Enter DOB (YYYY-MM-DD): ")
    while not validate_date(dob):
        dob = input("Invalid date format. Enter DOB (YYYY-MM-DD): ")
    phone = input("Enter Phone: ")
    while not validate_phone(phone):
        phone = input("Enter Phone: ")
    email = input("Enter Email: ")
    while not validate_email(email):
        email = input("Enter Email: ")
    designation = input("Enter Designation: ")
    salary = float(input("Enter Salary: "))
    employee = Employee(name, dob, phone, email, designation, salary)
    return employee


def main():
    """Main function to manage the employee management system."""
    employee_list = []
    while True:
        print("\nMenu:")
        print("1. Add Employee")
        print("2. Show Employee Details")
        print("3. Save to Text")
        print("4. Save to Excel")
        print("5. Save to PDF")
        print("6. Bulk Read from Excel")
        print("7. Exit")

        choice = input("Enter your choice: ")

        if choice == '1':
            employee_list.append(get_employee_details())
            print("Employee added successfully.")

        elif choice == '2':
            show_employee_details(employee_list)

        elif choice == '3':
            save_to_text(employee_list)
            print("Data saved to text file.")

        elif choice == '4':
            save_to_excel(employee_list)
            print("Data saved to Excel file.")

        elif choice == '5':
            save_to_pdf(employee_list)
            print("Data saved to PDF file.")

        elif choice == '6':
            employee_list = bulk_read_from_excel()
            print("Data Loaded from Excel file.")

        elif choice == '7':
            print("Exiting program.")
            break

        else:
            print("Invalid choice. Please enter a valid option.")


if __name__ == "__main__":
    main()
